import pandas as pd
import numpy as np
import warnings
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')
if not hasattr(pd.errors, 'SettingWithCopyWarning'):
    class SettingWithCopyWarning(UserWarning):
        pass
    pd.errors.SettingWithCopyWarning = SettingWithCopyWarning
    pd.core.common.SettingWithCopyWarning = SettingWithCopyWarning

from comorbidipy import comorbidity



CHARLSON_ICD10_MAPPING = {
    'CEVD': {
        'name': 'Cerebrovascular Disease',
        'codes': ['G45', 'G46', 'I60', 'I61', 'I62', 'I63', 'I64', 'I65', 'I66', 'I67', 'I68', 'I69'],
        'points': 1
    },
    'CHF': {
        'name': 'Congestive Heart Failure',
        'codes': ['I50'],
        'points': 1
    },
    'PVD': {
        'name': 'Peripheral Vascular Disease',
        'codes': ['I70', 'I71', 'I72', 'I73', 'I74', 'I77', 'I78', 'I79', 'K55'],
        'points': 1
    },
    'DEMENTIA': {
        'name': 'Dementia',
        'codes': ['F01', 'F02', 'F03', 'G30'],
        'points': 1
    },
    'COPD': {
        'name': 'Chronic Obstructive Pulmonary Disease',
        'codes': ['J41', 'J42', 'J43', 'J44', 'J45', 'J46', 'J47', 'J60', 'J61', 'J62', 'J63', 'J64', 'J65', 'J66', 'J67'],
        'points': 1
    },
    'RHEUMD': {
        'name': 'Rheumatologic Disease',
        'codes': ['M05', 'M06', 'M31', 'M32', 'M33', 'M34', 'M35', 'M36'],
        'points': 1
    },
    'PUD': {
        'name': 'Peptic Ulcer Disease',
        'codes': ['K25', 'K26', 'K27', 'K28'],
        'points': 1
    },
    'MLD': {
        'name': 'Mild Liver Disease',
        'codes': ['B18', 'C80', 'K70', 'K71', 'K73', 'K74', 'K76', 'Z94'],
        'points': 1
    },
    'DIAB': {
        'name': 'Diabetes (without complications)',
        'codes': ['E10', 'E11', 'E12', 'E13', 'E14'],
        'points': 1
    },
    'DIABWC': {
        'name': 'Diabetes with Complications',
        'codes': ['E10', 'E11', 'E12', 'E13', 'E14'],  # Combined with complications marker
        'points': 2
    },
    'REND': {
        'name': 'Renal Disease',
        'codes': ['I12', 'I13', 'N03', 'N05', 'N18', 'N19', 'N25', 'Z49'],
        'points': 2
    },
    'CANC': {
        'name': 'Cancer (non-metastatic)',
        'codes': ['C00', 'C01', 'C02', 'C03', 'C04', 'C05', 'C06', 'C07', 'C08', 'C09', 'C10', 'C11', 'C12', 'C13', 'C14', 'C15', 'C16', 'C17', 'C18', 'C19', 'C20', 'C21', 'C22', 'C23', 'C24', 'C25', 'C26', 'C30', 'C31', 'C32', 'C33', 'C34', 'C37', 'C38', 'C39', 'C40', 'C41', 'C43', 'C45', 'C47', 'C48', 'C49', 'C50', 'C51', 'C52', 'C53', 'C54', 'C55', 'C56', 'C57', 'C58', 'C60', 'C61', 'C62', 'C63', 'C64', 'C65', 'C66', 'C67', 'C68', 'C69', 'C70', 'C71', 'C72', 'C73', 'C74', 'C75', 'C76', 'C77', 'C78', 'C80', 'C81', 'C82', 'C83', 'C84', 'C85', 'C86', 'C87', 'C88', 'C89', 'C90', 'C91', 'C92', 'C93', 'C94', 'C95', 'C96', 'C97'],
        'points': 2
    },
    'METACANC': {
        'name': 'Metastatic Cancer',
        'codes': ['C77', 'C78', 'C79', 'C80'],
        'points': 6
    },
    'MSLD': {
        'name': 'Moderate/Severe Liver Disease',
        'codes': ['I85', 'I86', 'I87', 'K70', 'K71', 'K72', 'K73', 'K74'],
        'points': 3
    },
    'AIDS': {
        'name': 'AIDS/HIV',
        'codes': ['B20', 'B21', 'B22', 'B23', 'B24'],
        'points': 6
    }
}

class AlignedCharlsonCalculator:
    
    def __init__(self):
        self.mapping = CHARLSON_ICD10_MAPPING
    
    def extract_icd_codes(self, row):
        codes = []
        for i in range(1, 13):
            code_col = f'ICD_DGNS_CD{i}'
            if code_col in row.index and pd.notna(row[code_col]) and row[code_col] != '':
                codes.append(str(row[code_col]).strip().upper())
        return codes
    
    def check_condition(self, icd_codes, condition_codes):

        for icd in icd_codes:
            for cond_code in condition_codes:
                # Remove trailing dots for comparison
                prefix = cond_code.rstrip('.')
                if icd.startswith(prefix):
                    return True
        return False
    
    def calculate_cci(self, row):
        icd_codes = self.extract_icd_codes(row)
        
        conditions_present = {}
        score = 0
        
        # Check each condition
        for condition_key, condition_info in self.mapping.items():
            if self.check_condition(icd_codes, condition_info['codes']):
                conditions_present[condition_key] = 1
                score += condition_info['points']
            else:
                conditions_present[condition_key] = 0
        
        return score, conditions_present, icd_codes
    
    def process_dataframe(self, df):
        results = []
        
        for idx, row in df.iterrows():
            score, conditions, codes = self.calculate_cci(row)
            
            result = {
                'DSYSRTKY': row['DSYSRTKY'],
                'CLAIMNO': row['CLAIMNO'],
                'Aligned_CCI_Score': score,
                'ICD_Codes': '|'.join(codes),
            }
            result.update(conditions)
            results.append(result)
        
        return pd.DataFrame(results)

def calculate_comorbidipy_cci(df):
    try:
        comorbidity_data = []
        
        for idx, row in df.iterrows():
            patient_id = row['DSYSRTKY']
            
            try:
                if 'DOB_DT' in row.index and pd.notna(row['DOB_DT']):
                    dob = pd.to_datetime(row['DOB_DT'])
                    age = (datetime.now() - dob).days // 365
                else:
                    age = 65
            except:
                age = 65
            
            for i in range(1, 13):
                code_col = f'ICD_DGNS_CD{i}'
                if code_col in row.index and pd.notna(row[code_col]):
                    icd_code = str(row[code_col]).strip().upper()
                    if icd_code:
                        comorbidity_data.append({
                            'id': patient_id,
                            'code': icd_code,
                            'age': age
                        })
        
        comorbidity_df = pd.DataFrame(comorbidity_data)
        
        if len(comorbidity_df) == 0:
            return None
        
        # Calculate using comorbidipy
        result = comorbidity(
            comorbidity_df,
            id='id',
            code='code',
            age='age',
            score='charlson',
            icd='icd10',
            variant='quan',
            weighting='quan',
            assign0=True
        )
        
        result_clean = result[['id', 'comorbidity_score']].copy()
        result_clean.rename(columns={
            'id': 'DSYSRTKY',
            'comorbidity_score': 'Comorbidipy_CCI_Score'
        }, inplace=True)
        
        result_clean = result_clean.drop_duplicates(subset=['DSYSRTKY'], keep='first')
        
        return result_clean
        
    except Exception as e:
        print(f"   ❌ Comorbidipy error: {str(e)[:100]}")
        return None

def main():
    print("\n" + "="*80)
    print("ALIGNED CCI COMPARISON - CUSTOM CALCULATOR VS COMORBIDIPY")
    print("="*80 + "\n")
    
    # Load data
    print("1️⃣  Loading patient dataset...")
    df = pd.read_csv('synthetic_dmerc_base 1.csv')
    print(f"   ✅ Loaded {len(df)} patient records\n")
    
    # Calculate with aligned custom calculator
    print("2️⃣  Calculating with ALIGNED Custom Calculator (17 conditions)...")
    calculator = AlignedCharlsonCalculator()
    aligned_results = calculator.process_dataframe(df)
    print(f"   ✅ Calculated for {len(aligned_results)} patients\n")
    
    # Calculate with Comorbidipy
    print("3️⃣  Calculating with Comorbidipy...")
    comorbidipy_results = calculate_comorbidipy_cci(df)
    
    if comorbidipy_results is None:
        print("   ❌ Comorbidipy failed\n")
        return
    print(f"   ✅ Calculated for {len(comorbidipy_results)} patients\n")
    
    # Merge results
    print("4️⃣  Comparing results...")
    comparison = aligned_results[['DSYSRTKY', 'CLAIMNO', 'Aligned_CCI_Score']].copy()
    comparison = comparison.merge(
        comorbidipy_results,
        on='DSYSRTKY',
        how='left'
    )
    
    # Calculate agreement
    comparison['Match'] = comparison['Aligned_CCI_Score'] == comparison['Comorbidipy_CCI_Score']
    comparison['Difference'] = abs(comparison['Aligned_CCI_Score'] - comparison['Comorbidipy_CCI_Score'])
    
    matches = comparison['Match'].sum()
    total = len(comparison)
    agreement_rate = (matches / total * 100) if total > 0 else 0
    
    print(f"   ✅ Comparison complete\n")
    
    # Print statistics
    print("="*80)
    print("AGREEMENT ANALYSIS")
    print("="*80)
    print(f"\nTotal Patients: {total}")
    print(f"Exact Matches: {matches}/{total} ({agreement_rate:.1f}%)")
    print(f"Discrepancies: {total - matches}")
    
    if total - matches > 0:
        print(f"\nFor mismatches:")
        avg_diff = comparison[comparison['Match'] == False]['Difference'].mean()
        max_diff = comparison['Difference'].max()
        print(f"  Average Difference: {avg_diff:.2f}")
        print(f"  Max Difference: {max_diff:.0f}")
        
        # Show some examples of mismatches
        mismatches = comparison[comparison['Match'] == False].head(5)
        print(f"\nExample mismatches:")
        for _, row in mismatches.iterrows():
            print(f"  Patient {row['DSYSRTKY']}: Aligned={row['Aligned_CCI_Score']:.0f}, Comorbidipy={row['Comorbidipy_CCI_Score']:.0f}, Diff={row['Difference']:.0f}")
    
    print("\nScore Ranges:")
    print(f"  Aligned Custom: {aligned_results['Aligned_CCI_Score'].min():.0f} - {aligned_results['Aligned_CCI_Score'].max():.0f}")
    print(f"  Comorbidipy: {comorbidipy_results['Comorbidipy_CCI_Score'].min():.0f} - {comorbidipy_results['Comorbidipy_CCI_Score'].max():.0f}")
    
    print("\n" + "="*80 + "\n")
    
    # Create Excel file
    print("5️⃣  Creating comparison Excel file...")
    create_comparison_excel('CCI_Aligned_vs_Comorbidipy_100pct.xlsx', comparison)
    print("   ✅ Created: CCI_Aligned_vs_Comorbidipy_100pct.xlsx\n")
    
    print("="*80)
    print("✅ ANALYSIS COMPLETE!")
    print("="*80)
    print(f"\nAgreement Rate: {agreement_rate:.1f}%")
    print("\n")

def create_comparison_excel(filename, df):
    """Create professionally formatted comparison Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "Charlson CCI - Aligned Custom Calculator vs Comorbidipy (17 Conditions)"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Summary
    matches = (df['Match'] == True).sum()
    total = len(df)
    agreement_rate = (matches / total * 100) if total > 0 else 0
    
    ws['A2'] = f"Agreement Rate: {agreement_rate:.1f}% ({matches}/{total})"
    ws['A2'].font = Font(size=11, bold=True, color="FFFFFF")
    
    if agreement_rate >= 95:
        ws['A2'].fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")  # Green
    elif agreement_rate >= 80:
        ws['A2'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Yellow
    else:
        ws['A2'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
    
    ws.merge_cells('A2:H2')
    
    ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'].font = Font(size=10, italic=True, color="666666")
    
    # Headers
    start_row = 5
    columns = list(df.columns)
    
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = col_name
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.row_dimensions[start_row].height = 30
    
    # Data rows
    for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            # Alternate row colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
            
            # Color code Match column
            if columns[col_idx - 1] == 'Match':
                if value == True:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100", bold=True)
                    cell.value = "✓ MATCH"
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006", bold=True)
                    cell.value = "✗ DIFF"
            
            # Format alignment
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
    
    # Adjust column widths
    for col_idx, col_name in enumerate(columns, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 15
    
    # Freeze panes
    ws.freeze_panes = f'A{start_row + 1}'
    
    wb.save(filename)

if __name__ == '__main__':
    main()
