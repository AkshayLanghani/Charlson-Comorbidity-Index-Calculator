import pandas as pd
import numpy as np
import warnings
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')
if not hasattr(pd.errors, 'SettingWithCopyWarning'):
    class SettingWithCopyWarning(UserWarning):
        pass
    pd.errors.SettingWithCopyWarning = SettingWithCopyWarning
    pd.core.common.SettingWithCopyWarning = SettingWithCopyWarning

from comorbidipy import comorbidity

# ============================================================================
# CUSTOM CCI CALCULATOR (17 CONDITIONS)
# ============================================================================

CHARLSON_CONDITIONS = {
    'CEVD': {'name': 'Cerebrovascular Disease', 'codes': ['G45', 'G46', 'I60', 'I61', 'I62', 'I63', 'I64', 'I65', 'I66', 'I67', 'I68', 'I69'], 'points': 1},
    'CHF': {'name': 'Congestive Heart Failure', 'codes': ['I50'], 'points': 1},
    'PVD': {'name': 'Peripheral Vascular Disease', 'codes': ['I70', 'I71', 'I72', 'I73', 'I74', 'I77', 'I78', 'I79', 'K55'], 'points': 1},
    'DEMENTIA': {'name': 'Dementia', 'codes': ['F01', 'F02', 'F03', 'G30'], 'points': 1},
    'COPD': {'name': 'COPD', 'codes': ['J41', 'J42', 'J43', 'J44', 'J45', 'J46', 'J47', 'J60', 'J61', 'J62', 'J63', 'J64', 'J65', 'J66', 'J67'], 'points': 1},
    'RHEUMD': {'name': 'Rheumatologic', 'codes': ['M05', 'M06', 'M31', 'M32', 'M33', 'M34', 'M35', 'M36'], 'points': 1},
    'PUD': {'name': 'Peptic Ulcer', 'codes': ['K25', 'K26', 'K27', 'K28'], 'points': 1},
    'MLD': {'name': 'Mild Liver Dis', 'codes': ['B18', 'C80', 'K70', 'K71', 'K73', 'K74', 'K76', 'Z94'], 'points': 1},
    'DIAB': {'name': 'Diabetes', 'codes': ['E10', 'E11', 'E12', 'E13', 'E14'], 'points': 1},
    'REND': {'name': 'Renal Disease', 'codes': ['I12', 'I13', 'N03', 'N05', 'N18', 'N19', 'N25', 'Z49'], 'points': 2},
    'CANC': {'name': 'Cancer (non-met)', 'codes': ['C00', 'C01', 'C02', 'C03', 'C04', 'C05', 'C06', 'C07', 'C08', 'C09', 'C10', 'C11', 'C12', 'C13', 'C14', 'C15', 'C16', 'C17', 'C18', 'C19', 'C20', 'C21', 'C22', 'C23', 'C24', 'C25', 'C26', 'C30', 'C31', 'C32', 'C33', 'C34', 'C37', 'C38', 'C39', 'C40', 'C41', 'C43', 'C45', 'C47', 'C48', 'C49', 'C50', 'C51', 'C52', 'C53', 'C54', 'C55', 'C56', 'C57', 'C58', 'C60', 'C61', 'C62', 'C63', 'C64', 'C65', 'C66', 'C67', 'C68', 'C69', 'C70', 'C71', 'C72', 'C73', 'C74', 'C75', 'C76', 'C77', 'C78', 'C80', 'C81', 'C82', 'C83', 'C84', 'C85', 'C86', 'C87', 'C88', 'C89', 'C90', 'C91', 'C92', 'C93', 'C94', 'C95', 'C96', 'C97'], 'points': 2},
    'METACANC': {'name': 'Metastatic CA', 'codes': ['C77', 'C78', 'C79', 'C80'], 'points': 6},
    'MSLD': {'name': 'Severe Liver', 'codes': ['I85', 'I86', 'I87', 'K70', 'K71', 'K72', 'K73', 'K74'], 'points': 3},
    'AIDS': {'name': 'AIDS', 'codes': ['B20', 'B21', 'B22', 'B23', 'B24'], 'points': 6}
}

class CustomCharlsonCalculator:
    def __init__(self):
        self.conditions = CHARLSON_CONDITIONS
    
    def extract_codes(self, row):
        codes = []
        for i in range(1, 13):
            col = f'ICD_DGNS_CD{i}'
            if col in row.index and pd.notna(row[col]):
                codes.append(str(row[col]).strip().upper())
        return codes
    
    def check_condition(self, codes, condition_codes):
        for code in codes:
            for cond_code in condition_codes:
                if code.startswith(cond_code.rstrip('.')):
                    return True
        return False
    
    def calculate(self, row):
        codes = self.extract_codes(row)
        has_codes = len(codes) > 0
        
        if not has_codes:
            return 0, {}, codes, False
        
        conditions = {}
        score = 0
        for cond_key, cond_info in self.conditions.items():
            if self.check_condition(codes, cond_info['codes']):
                conditions[cond_key] = 1
                score += cond_info['points']
            else:
                conditions[cond_key] = 0
        
        return score, conditions, codes, True

def process_custom_calculator(df):
    """Process all 839 patients with custom calculator"""
    calc = CustomCharlsonCalculator()
    results = []
    
    for idx, row in df.iterrows():
        score, conditions, codes, has_codes = calc.calculate(row)
        
        result = {
            'DSYSRTKY': row['DSYSRTKY'],
            'CLAIMNO': row['CLAIMNO'],
            'Custom_CCI_Score': score if has_codes else np.nan,
            'Has_ICD_Codes': 'Yes' if has_codes else 'No',
            'ICD_Codes': '|'.join(codes) if codes else 'None',
        }
        result.update(conditions)
        results.append(result)
    
    return pd.DataFrame(results)

def process_comorbidipy(df):
    """Process patients with comorbidipy"""
    data = []
    
    for idx, row in df.iterrows():
        patient_id = row['DSYSRTKY']
        try:
            dob = pd.to_datetime(row['DOB_DT']) if 'DOB_DT' in row.index and pd.notna(row['DOB_DT']) else None
            age = (datetime.now() - dob).days // 365 if dob else 65
        except:
            age = 65
        
        for i in range(1, 13):
            col = f'ICD_DGNS_CD{i}'
            if col in row.index and pd.notna(row[col]):
                code = str(row[col]).strip().upper()
                if code:
                    data.append({'id': patient_id, 'code': code, 'age': age})
    
    if len(data) == 0:
        return None
    
    df_data = pd.DataFrame(data)
    result = comorbidity(df_data, id='id', code='code', age='age', score='charlson', icd='icd10')
    result_clean = result[['id', 'comorbidity_score']].copy()
    result_clean.rename(columns={'id': 'DSYSRTKY', 'comorbidity_score': 'Comorbidipy_CCI_Score'}, inplace=True)
    result_clean = result_clean.drop_duplicates(subset=['DSYSRTKY'], keep='first')
    
    return result_clean

def main():
    print("\n" + "="*80)
    print("COMPREHENSIVE CCI ANALYSIS - ALL 839 PATIENTS")
    print("="*80 + "\n")
    
    # Load data
    print("1️⃣  Loading dataset...")
    df = pd.read_csv('synthetic_dmerc_base 1.csv')
    print(f"   ✅ Loaded {len(df)} patient records\n")
    
    # Custom calculator
    print("2️⃣  Calculating with Custom Calculator (17 conditions)...")
    custom_df = process_custom_calculator(df)
    has_codes = (custom_df['Has_ICD_Codes'] == 'Yes').sum()
    print(f"   ✅ {has_codes}/839 patients have ICD codes\n")
    
    # Comorbidipy
    print("3️⃣  Calculating with Comorbidipy...")
    combo_df = process_comorbidipy(df)
    if combo_df is not None:
        print(f"   ✅ {len(combo_df)} patients scored\n")
    
    # Merge all
    print("4️⃣  Merging results...")
    all_results = custom_df.copy()
    if combo_df is not None:
        all_results = all_results.merge(combo_df, on='DSYSRTKY', how='left')
    
    all_results['Match'] = (all_results['Custom_CCI_Score'] == all_results['Comorbidipy_CCI_Score'])
    
    # Create Excel workbook
    print("5️⃣  Creating professional Excel workbook...")
    create_comprehensive_excel('CCI_Complete_Analysis_839_Patients.xlsx', df, all_results, custom_df, combo_df)
    
    print("\n" + "="*80)
    print("✅ ANALYSIS COMPLETE!")
    print("="*80)
    print(f"\nFile: CCI_Complete_Analysis_839_Patients.xlsx")
    print(f"Total Patients: {len(df)}")
    print(f"Patients with ICD codes: {has_codes}")
    print(f"Patients scored by Comorbidipy: {len(combo_df) if combo_df is not None else 0}")
    print("\n")

def create_comprehensive_excel(filename, raw_df, all_results, custom_df, combo_df):
    """Create comprehensive multi-sheet Excel workbook"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Sheet 1: All Patients with CCI Scores
    ws1 = wb.create_sheet("CCI Results (All 839)", 0)
    add_cci_results_sheet(ws1, all_results)
    
    # Sheet 2: Custom Calculator Details
    ws2 = wb.create_sheet("Custom Calculator")
    add_custom_details_sheet(ws2, custom_df)
    
    # Sheet 3: Comorbidipy Results
    if combo_df is not None:
        ws3 = wb.create_sheet("Comorbidipy Results")
        add_comorbidipy_sheet(ws3, combo_df)
    
    # Sheet 4: Comparison
    ws4 = wb.create_sheet("Comparison")
    add_comparison_sheet(ws4, all_results)
    
    # Sheet 5: Demographics & Summary
    ws5 = wb.create_sheet("Summary Statistics")
    add_summary_sheet(ws5, raw_df, all_results, custom_df, combo_df)
    
    # Sheet 6: Condition Prevalence
    ws6 = wb.create_sheet("Condition Prevalence")
    add_prevalence_sheet(ws6, custom_df)
    
    wb.save(filename)

def add_cci_results_sheet(ws, df):
    """Sheet 1: All patient results"""
    # Title and info
    ws['A1'] = "CCI Analysis - All 839 Patients"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws.merge_cells('A1:G1')
    ws.row_dimensions[1].height = 25
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(size=10, italic=True, color="666666")
    ws['A3'] = f"Total Records: {len(df)}"
    ws['A3'].font = Font(size=10, italic=True, color="666666")
    
    # Headers
    headers = ['DSYSRTKY', 'CLAIMNO', 'Custom_CCI', 'Comorbidipy_CCI', 'Match', 'Has_ICD', 'ICD_Codes']
    start_row = 5
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[start_row].height = 25
    
    # Data
    for row_idx, row_data in enumerate(dataframe_to_rows(df[['DSYSRTKY', 'CLAIMNO', 'Custom_CCI_Score', 'Comorbidipy_CCI_Score', 'Match', 'Has_ICD_Codes', 'ICD_Codes']], index=False, header=False), start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            # Alternating colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="F2F8FF", end_color="F2F8FF", fill_type="solid")
            
            # Color code Match column
            if headers[col_idx - 1] == 'Match':
                if value == True:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(bold=True, color="006100")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(bold=True, color="9C0006")
            
            # Color code Has_ICD
            if headers[col_idx - 1] == 'Has_ICD':
                if value == 'Yes':
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                    cell.font = Font(color="155724")
                else:
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                    cell.font = Font(color="721C24")
            
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Adjust widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 40
    
    ws.freeze_panes = f'A{start_row + 1}'

def add_custom_details_sheet(ws, df):
    """Sheet 2: Custom calculator details"""
    ws['A1'] = "Custom CCI Calculator - Detailed Results"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 25
    
    # Summary stats
    ws['A3'] = "Summary:"
    ws['A4'] = f"Total Patients: {len(df)}"
    ws['A5'] = f"Patients with ICD Codes: {(df['Has_ICD_Codes'] == 'Yes').sum()}"
    ws['A6'] = f"Mean CCI Score: {df['Custom_CCI_Score'].mean():.2f}"
    ws['A7'] = f"Median CCI Score: {df['Custom_CCI_Score'].median():.0f}"
    
    # Headers
    headers = list(df.columns)
    start_row = 9
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Data
    for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col_idx in range(1, len(headers) + 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

def add_comorbidipy_sheet(ws, df):
    """Sheet 3: Comorbidipy results"""
    ws['A1'] = "Comorbidipy CCI Results"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 25
    
    ws['A3'] = f"Patients Scored: {len(df)}"
    ws['A4'] = f"Mean Score: {df['Comorbidipy_CCI_Score'].mean():.2f}"
    
    headers = ['DSYSRTKY', 'Comorbidipy_CCI_Score']
    start_row = 6
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20

def add_comparison_sheet(ws, df):
    """Sheet 4: Comparison"""
    ws['A1'] = "Custom vs Comorbidipy Comparison"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 25
    
    matched = df['Match'].sum()
    total = len(df[df['Has_ICD_Codes'] == 'Yes'])
    ws['A3'] = f"Agreement Rate: {(matched/total*100 if total > 0 else 0):.1f}% ({matched}/{total})"
    
    headers = ['DSYSRTKY', 'Custom_CCI_Score', 'Comorbidipy_CCI_Score', 'Match']
    start_row = 5
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_idx, row_data in enumerate(dataframe_to_rows(df[['DSYSRTKY', 'Custom_CCI_Score', 'Comorbidipy_CCI_Score', 'Match']], index=False, header=False), start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            if headers[col_idx - 1] == 'Match':
                if value == True:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(bold=True, color="006100")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(bold=True, color="9C0006")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 15
    ws.freeze_panes = f'A{start_row + 1}'

def add_summary_sheet(ws, raw_df, all_results, custom_df, combo_df):
    """Sheet 5: Summary statistics"""
    ws['A1'] = "Summary Statistics & Demographics"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    ws.merge_cells('A1:B1')
    ws.row_dimensions[1].height = 25
    
    row = 3
    metrics = [
        ('Total Patients in Dataset', len(raw_df)),
        ('Patients with ICD Codes', (custom_df['Has_ICD_Codes'] == 'Yes').sum()),
        ('Patients without ICD Codes', (custom_df['Has_ICD_Codes'] == 'No').sum()),
        ('Patients Scored by Comorbidipy', len(combo_df) if combo_df is not None else 0),
        ('', ''),
        ('CUSTOM CALCULATOR STATISTICS', ''),
        ('Mean CCI Score', f"{custom_df['Custom_CCI_Score'].mean():.2f}"),
        ('Median CCI Score', f"{custom_df['Custom_CCI_Score'].median():.0f}"),
        ('Min Score', f"{custom_df['Custom_CCI_Score'].min():.0f}"),
        ('Max Score', f"{custom_df['Custom_CCI_Score'].max():.0f}"),
        ('', ''),
        ('COMORBIDIPY STATISTICS', ''),
        ('Mean CCI Score', f"{combo_df['Comorbidipy_CCI_Score'].mean():.2f}" if combo_df is not None else "N/A"),
        ('Median CCI Score', f"{combo_df['Comorbidipy_CCI_Score'].median():.0f}" if combo_df is not None else "N/A"),
        ('Min Score', f"{combo_df['Comorbidipy_CCI_Score'].min():.0f}" if combo_df is not None else "N/A"),
        ('Max Score', f"{combo_df['Comorbidipy_CCI_Score'].max():.0f}" if combo_df is not None else "N/A"),
    ]
    
    for label, value in metrics:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        if label.isupper() and label != '':
            ws[f'A{row}'].fill = PatternFill(start_color="E6D9F2", end_color="E6D9F2", fill_type="solid")
            ws[f'B{row}'].fill = PatternFill(start_color="E6D9F2", end_color="E6D9F2", fill_type="solid")
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
        row += 1
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

def add_prevalence_sheet(ws, df):
    """Sheet 6: Condition prevalence"""
    ws['A1'] = "Condition Prevalence Analysis"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 25
    
    condition_cols = [c for c in df.columns if c not in ['DSYSRTKY', 'CLAIMNO', 'Custom_CCI_Score', 'Has_ICD_Codes', 'ICD_Codes']]
    
    headers = ['Condition', 'Count', 'Percentage', 'CCI Points']
    start_row = 3
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="FF8787", end_color="FF8787", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row = start_row + 1
    for cond in sorted(condition_cols):
        count = df[cond].sum()
        pct = (count / len(df) * 100) if len(df) > 0 else 0
        
        ws[f'A{row}'] = cond
        ws[f'B{row}'] = int(count)
        ws[f'C{row}'] = f"{pct:.1f}%"
        ws[f'D{row}'] = ""
        
        if row % 2 == 0:
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{col}{row}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        row += 1
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12

if __name__ == '__main__':
    main()
